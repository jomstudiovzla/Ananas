import os
import time
import requests
from bs4 import BeautifulSoup
import openpyxl
from PIL import Image
from rembg import remove
import urllib.parse
from sync_excel_to_db import sync

# Optimized search query map for the 83 catalog products
QUERY_MAP = {
    "p1": "tomate perita",
    "p2": "cebolla blanca",
    "p3": "papa",
    "p4": "zanahoria",
    "p5": "pimenton verde",
    "p6": "lechuga romana",
    "p7": "platano",
    "p8": "cambur",
    "p9": "manzana gala",
    "p10": "naranja",
    "p11": "lechosa",
    "p12": "limon",
    "p13": "piña",
    "p14": "aguacate",
    "p15": "patilla",
    "p16": "carne molida",
    "p17": "bistec",
    "p18": "costilla res",
    "p19": "pechuga pollo",
    "p20": "pollo entero",
    "p21": "muslos pollo",
    "p22": "alitas pollo",
    "p23": "jamon plumrose",
    "p24": "queso amarillo paisa",
    "p25": "queso blanco",
    "p26": "salchicha plumrose",
    "p27": "tocineta",
    "p28": "chorizo",
    "p29": "queso guayanes",
    "p30": "nuggets",
    "p31": "coca cola 2l",
    "p32": "pepsi 2l",
    "p33": "chinotto 2l",
    "p34": "frescolita 2l",
    "p35": "maltin 1.5",
    "p36": "arroz mary",
    "p37": "arroz primor",
    "p38": "pasta capri",
    "p39": "harina pan",
    "p40": "caraotas mary",
    "p41": "lentejas",
    "p42": "atun margarita aceite",
    "p43": "atun margarita agua",
    "p44": "maiz dulce",
    "p45": "mazeite",
    "p46": "protex",
    "p47": "dove",
    "p48": "colgate",
    "p49": "pantene",
    "p50": "scott",
    "p51": "always",
    "p52": "gillette",
    "p53": "lubriderm",
    "p54": "algodon",
    "p55": "hisopos",
    "p56": "alcohol",
    "p57": "agua oxigenada",
    "p58": "curitas",
    "p59": "ariel liquido",
    "p60": "ace polvo",
    "p61": "suavitel",
    "p62": "las llaves lavaplatos",
    "p63": "axion crema",
    "p64": "cloro nevex",
    "p65": "mistolin",
    "p66": "esponja scotch",
    "p67": "mopa",
    "p68": "escoba",
    "p69": "coleto",
    "p70": "bolsas basura",
    "p71": "polar pilsen",
    "p72": "polar light",
    "p73": "solera",
    "p74": "zulia",
    "p75": "santa teresa linaje",
    "p76": "cacique anejo",
    "p77": "pampero aniversario",
    "p78": "diplomatico reserva",
    "p79": "buchanans",
    "p80": "old parr",
    "p81": "gordons",
    "p82": "casillero del diablo",
    "p83": "santa helena"
}

def clean_price(price_str):
    try:
        # Format: "$ 1,44" or "Ref: 3.45" -> parse to float
        price_str = price_str.replace('$', '').replace('Ref:', '').replace('ref:', '').strip()
        price_str = price_str.replace(',', '.')
        return float(price_str)
    except Exception:
        return None

def main():
    excel_path = "public/data/productos_inventario.xlsx"
    if not os.path.exists(excel_path):
        print(f"Error: {excel_path} no existe.")
        return
        
    wb = openpyxl.load_workbook(excel_path)
    ws = wb.active
    
    headers = {
        "User-Agent": "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36"
    }
    
    print("Iniciando la actualización masiva de inventario con datos reales de Automercados Plaza's...")
    
    # Process each product row
    for r in range(2, ws.max_row + 1):
        pid = ws.cell(row=r, column=1).value
        orig_name = ws.cell(row=r, column=2).value
        image_file = ws.cell(row=r, column=7).value
        
        if not pid or not orig_name:
            continue
            
        pid_str = str(pid).strip()
        query = QUERY_MAP.get(pid_str, orig_name.lower())
        
        print(f"\n[{pid_str}] Buscando '{query}' (original: {orig_name})...")
        encoded_query = urllib.parse.quote_plus(query)
        search_url = f"https://vallearriba.elplazas.com/catalogsearch/result/?q={encoded_query}"
        
        try:
            res = requests.get(search_url, headers=headers, timeout=15)
            if res.status_code != 200:
                print(f"  ⚠ Error de red {res.status_code}. Pasando al siguiente.")
                continue
                
            soup = BeautifulSoup(res.text, 'html.parser')
            products = soup.find_all(class_='product-item')
            
            if not products:
                print(f"  ⚠ No se encontraron resultados para '{query}' en El Plazas.")
                continue
                
            # Take the first matched product card
            first_product = products[0]
            
            name_tag = first_product.find(class_='product-item-link')
            real_name = name_tag.text.strip() if name_tag else None
            
            price_tag = first_product.find(class_='price')
            real_price_val = price_tag.text.strip() if price_tag else None
            
            img_tag = first_product.find('img')
            img_url = None
            if img_tag:
                img_url = img_tag.get('src') or img_tag.get('data-src') or img_tag.get('data-original')
            
            if not real_name or not real_price_val or not img_url:
                print("  ⚠ Datos del producto incompletos en el HTML.")
                continue
                
            parsed_price = clean_price(real_price_val)
            if parsed_price is None:
                print(f"  ⚠ No se pudo parsear el precio: {real_price_val}")
                continue
                
            print(f"  ✓ Encontrado: '{real_name}'")
            print(f"  ✓ Precio Real Plazas: ${parsed_price:.2f} (original: {real_price_val})")
            print(f"  ✓ URL de imagen: {img_url}")
            
            # Update Excel spreadsheet data
            # Update Name
            ws.cell(row=r, column=2, value=real_name)
            # Update Price
            ws.cell(row=r, column=3, value=parsed_price)
            
            # Download and clean the image
            print("  Descargando y procesando la imagen...")
            img_res = requests.get(img_url, headers=headers, timeout=15)
            if img_res.status_code == 200:
                # Save to a temp file
                temp_img_path = f"temp_{pid_str}.jpg"
                with open(temp_img_path, "wb") as f_img:
                    f_img.write(img_res.content)
                    
                # Clean background
                img_obj = Image.open(temp_img_path).convert("RGBA")
                no_bg = remove(img_obj)
                
                # solid white bg
                white_bg = Image.new("RGBA", no_bg.size, (255, 255, 255, 255))
                combined = Image.alpha_composite(white_bg, no_bg)
                final = combined.convert("RGB")
                
                # Make sure filename ends with .png
                if not image_file:
                    image_file = f"{pid_str}.png"
                elif not str(image_file).endswith('.png'):
                    image_file = str(image_file).split('.')[0] + '.png'
                
                final_dest_path = f"public/images/products/{image_file}"
                final.save(final_dest_path, "PNG")
                print(f"  ✓ Imagen procesada y guardada en {final_dest_path}")
                
                # Clean temp file
                if os.path.exists(temp_img_path):
                    os.remove(temp_img_path)
            else:
                print(f"  ⚠ Error al descargar imagen: {img_res.status_code}")
                
        except Exception as e:
            print(f"  ✗ Error al procesar {pid_str}: {e}")
            
        # Small delay to respect rate limiting
        time.sleep(0.5)
        
    print("Guardando el archivo Excel final...")
    wb.save(excel_path)
    print("\n--- Scraping finalizado. Iniciando sincronización de base de datos... ---")
    sync()
    print("--- ¡Sincronización completa! ---")

if __name__ == "__main__":
    main()
