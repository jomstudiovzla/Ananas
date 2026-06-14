import openpyxl
import os
from sync_excel_to_db import sync

def patch():
    excel_path = "public/data/productos_inventario.xlsx"
    if not os.path.exists(excel_path):
        print(f"Error: {excel_path} no existe.")
        return
        
    wb = openpyxl.load_workbook(excel_path)
    ws = wb.active
    
    # We will map updates by ID
    # Column mappings:
    # Col 1: ID
    # Col 2: Name
    # Col 3: Price
    # Col 4: Category
    # Col 5: Subcategory
    # Col 7: Image File
    # Col 9: Labels
    
    patches = {
        "p2": {
            "name": "Cebolla Blanca X KG",
            "price": 2.10,
            "subcategory": "Frescos",
            "image": "cebolla_blanca.png"
        },
        "p6": {
            "name": "Lechuga Romana",
            "price": 1.20,
            "subcategory": "Frescos",
            "image": "lechuga_romana.png"
        },
        "p20": {
            "name": "Pollo Entero X KG",
            "price": 4.50,
            "subcategory": "Pollos",
            "image": "pollo_entero.png"
        },
        "p22": {
            "name": "Alitas de Pollo X KG",
            "price": 6.00,
            "subcategory": "Pollos",
            "image": "alitas_pollo.png"
        },
        "p27": {
            "name": "Tocino Ahumado Plumrose 250G",
            "price": 4.80,
            "subcategory": "Embutidos",
            "image": "tocino.png"
        },
        "p39": {
            "subcategory": "Harinas"
        },
        "p40": {
            "subcategory": "Granos"
        },
        "p41": {
            "subcategory": "Granos"
        },
        "p45": {
            "subcategory": "Aceites"
        },
        "p74": {
            "name": "Cerveza Zulia 6 Pack",
            "price": 6.00,
            "subcategory": "Cervezas",
            "image": "cerveza_zulia.png"
        },
        "p75": {
            "name": "Ron Santa Teresa Linaje 750ML",
            "price": 18.00,
            "subcategory": "Destilados",
            "image": "ron_teresa.png"
        },
        "p77": {
            "name": "Ron Pampero Aniversario 750ML",
            "price": 22.00,
            "subcategory": "Destilados",
            "image": "ron_pampero.png"
        },
        "p82": {
            "name": "Vino Tinto Casillero del Diablo 750ML",
            "price": 12.50,
            "subcategory": "Vinos",
            "image": "vino_tinto.png"
        },
        "p83": {
            "name": "Vino Blanco Santa Helena 750ML",
            "price": 8.50,
            "subcategory": "Vinos",
            "image": "vino_blanco.png"
        }
    }
    
    print("Aplicando correcciones al archivo Excel...")
    for r in range(2, ws.max_row + 1):
        pid = ws.cell(row=r, column=1).value
        if not pid:
            continue
        pid_str = str(pid).strip()
        
        if pid_str in patches:
            p_data = patches[pid_str]
            if "name" in p_data:
                ws.cell(row=r, column=2, value=p_data["name"])
            if "price" in p_data:
                ws.cell(row=r, column=3, value=p_data["price"])
            if "subcategory" in p_data:
                ws.cell(row=r, column=5, value=p_data["subcategory"])
            if "image" in p_data:
                ws.cell(row=r, column=7, value=p_data["image"])
            print(f"  ✓ Producto {pid_str} parchado exitosamente.")
            
    wb.save(excel_path)
    print("Excel guardado. Sincronizando catálogo con base de datos web...")
    sync()
    print("¡Catálogo sincronizado exitosamente!")

if __name__ == "__main__":
    patch()
